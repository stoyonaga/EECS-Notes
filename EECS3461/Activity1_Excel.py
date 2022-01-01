from openpyxl import Workbook, load_workbook

# Variable Declarations
excel = load_workbook("Dataset6.xlsx")
defaultSheet = excel["dataset6"]
uniqueSheet = None
unique_entries = {}

# Generate Unique Entries Page for Tracking
if "Unique Entries" not in excel.sheetnames:
    excel.create_sheet("Unique Entries")
    uniqueSheet = excel["Unique Entries"]
    uniqueSheet.title = "Unique Entries"
    uniqueSheet["C1"] = "Additional Comments"
else:
    uniqueSheet = excel["Unique Entries"]
    print("WARNING: No need to generated sheet. It is already in the xlsx file!")

# Generate Dictionary of <Key = post title, Value = Empty List>
for row in range(1, 201):
    unique_entries[defaultSheet["F" + str(row)].value] = list()

# Populate Map-List with Duplicate DE-ID Value(s) for bookkeeping
for row in range(1, 201):
    unique_entries[defaultSheet["F" + str(row)].value].append(defaultSheet["A" + str(row)].value)

# Write (Key,Value) Pairs into the uniqueSheet
counter = 1
for key in unique_entries:
    uniqueSheet["A" + str(counter)] = key
    uniqueSheet["B" + str(counter)] = str(unique_entries[key])
    counter += 1

excel.save("Dataset6.xlsx")
