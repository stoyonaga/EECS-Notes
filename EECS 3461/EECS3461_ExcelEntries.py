from openpyxl import load_workbook


class Entry:
    # Class Entry will be used for storing key entries to write into the new page of our datasheet
    def __init__(self, link: str, ids: list):
        self.link = link
        self.ids = ids


# Variable Declarations
excel = load_workbook("Dataset6.xlsx")
defaultSheet = excel["dataset6"]
uniqueSheet = None
uniqueEntries = {}

# Generate and append a new page to the .xlsx file
if "Unique Entries" not in excel.sheetnames:
    excel.create_sheet("Unique Entries")
    uniqueSheet = excel["Unique Entries"]
    uniqueSheet.title = "Unique Entries"
    uniqueSheet["C1"] = "Classification"
    uniqueSheet["E1"] = "Additional Comments"

    # Determine the number of unique rows to generate in uniqueSheet
    for entry in range(1, 201):
        if defaultSheet["F" + str(entry)].value not in uniqueEntries.keys():
            newEntry = Entry(defaultSheet["E" + str(entry)].value, list())
            newEntry.ids.append(defaultSheet["A" + str(entry)].value)
            uniqueEntries[defaultSheet["F" + str(entry)].value] = newEntry
        else:
            uniqueEntries[defaultSheet["F" + str(entry)].value].ids.append(defaultSheet["A" + str(entry)].value)

    # Write the unique rows into the Unique Entries Page
    counter = 1
    for key in uniqueEntries.keys():
        uniqueSheet["A" + str(counter)] = str(key)
        uniqueSheet["B" + str(counter)] = str((uniqueEntries[key]).link)
        uniqueSheet["D" + str(counter)] = str((uniqueEntries[key]).ids)
        counter += 1
    uniqueSheet.column_dimensions["B"].width = 24
    uniqueSheet.column_dimensions["C"].width = 21.27
    uniqueSheet.column_dimensions["D"].width = 20
    uniqueSheet.column_dimensions["E"].width = 24
    excel.save("Dataset6.xlsx")
    print("Unique Entries sheet has been created.")
else:
    # Do not generate the page, it has already been generated before!
    print("Sheet has already been created. Check the Datasheet.xlsx file again.")
    uniqueSheet = excel["Unique Entries"]
